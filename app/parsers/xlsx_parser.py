from io import BytesIO
from openpyxl import load_workbook


def parse_xlsx(file_bytes: bytes) -> tuple[list[dict], int]:
    """
    Возвращает:
    sheets_data, sheet_count

    sheets_data = [
        {
            "sheet_name": str,
            "text": str,
            "rows": list[str]
        },
        ...
    ]
    """
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    sheets_data = []

    for sheet in workbook.worksheets:
        row_texts = []

        for row in sheet.iter_rows(values_only=True):
            values = []
            for cell in row:
                if cell is None:
                    continue

                value = str(cell).strip()
                if value:
                    values.append(value)

            if values:
                row_texts.append(" | ".join(values))

        sheet_text = "\n".join(row_texts).strip()

        sheets_data.append(
            {
                "sheet_name": sheet.title,
                "text": sheet_text,
                "rows": row_texts,
            }
        )

    return sheets_data, len(sheets_data)